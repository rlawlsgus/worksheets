# backend/routes/worklog_routes.py
from flask import Blueprint, request, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from sqlalchemy import extract
from io import BytesIO
from datetime import datetime, date
from models.worklog import WorkLog
from models.assistant import Assistant
from routes.auth_routes import is_admin, get_current_user_id, admin_required
from utils.time_calculator import calculate_work_hours
from database.db import db


worklog_bp = Blueprint("worklog", __name__)


@worklog_bp.route("/api/worklog", methods=["POST"])
def add_worklog():
    data = request.json
    work_hours = calculate_work_hours(data["start_time"], data["end_time"])
    start_datetime = datetime.strptime(data["start_time"], "%Y-%m-%d %H:%M")
    end_datetime = datetime.strptime(data["end_time"], "%Y-%m-%d %H:%M")

    if start_datetime.minute not in [0, 30] or end_datetime.minute not in [0, 30]:
        return (
            jsonify(
                {
                    "success": False,
                    "message": "Start time and end time must be on the hour or half-hour.",
                }
            ),
            400,
        )

    worklog = WorkLog(
        assistant_id=data["assistant_id"],
        date=start_datetime.date(),
        start_time=start_datetime.time(),
        end_time=end_datetime.time(),
        work_hours=work_hours,
    )
    db.session.add(worklog)
    db.session.commit()
    return jsonify({"success": True, "id": worklog.id})


@worklog_bp.route(
    "/api/worklogs/<int:assistant_id>/<int:year>/<int:month>", methods=["GET"]
)
def get_monthly_worklogs(assistant_id, year, month):
    if not is_admin() and assistant_id != get_current_user_id():
        return jsonify({"message": "Forbidden"}), 403

    worklogs = (
        WorkLog.query.filter(
            WorkLog.assistant_id == assistant_id,
            db.extract("year", WorkLog.date) == year,
            db.extract("month", WorkLog.date) == month,
        )
        .order_by(WorkLog.date)
        .all()
    )

    return jsonify(
        [
            {
                "id": log.id,
                "date": log.date.strftime("%Y-%m-%d"),
                "start_time": log.start_time.strftime("%H:%M"),
                "end_time": log.end_time.strftime("%H:%M"),
                "work_hours": log.work_hours,
                "checked": log.checked,
            }
            for log in worklogs
        ]
    )


@worklog_bp.route("/api/worklog/<int:id>", methods=["DELETE"])
def delete_worklog(id):
    worklog = WorkLog.query.get_or_404(id)
    db.session.delete(worklog)
    db.session.commit()
    return jsonify({"success": True})


@worklog_bp.route("/api/worklogs/check", methods=["PUT"])
@admin_required
def check_worklog():
    changed_rows = request.json
    for row in changed_rows:
        log = WorkLog.query.get(row["id"])
        if log:
            log.checked = row["checked"]
    db.session.commit()
    return jsonify({"success": True})


@worklog_bp.route("/api/worklogs/unchecked/<int:year>/<int:month>", methods=["GET"])
@admin_required
def get_unchecked_logs(year, month):
    # 해당 연도와 월에 대한 승인되지 않은 worklog 조회
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1)
    else:
        end_date = date(year, month + 1, 1)

    # 해당 월의 미승인 로그를 조회하고 assistant 이름도 함께 가져옴
    unchecked_logs = (
        db.session.query(WorkLog, Assistant.name.label("assistant_name"))
        .join(Assistant, WorkLog.assistant_id == Assistant.id)
        .filter(
            WorkLog.date >= start_date,
            WorkLog.date < end_date,
            WorkLog.checked == False,
        )
        .all()
    )

    result = [
        {**log.to_dict(), "assistant_name": assistant_name}
        for log, assistant_name in unchecked_logs
    ]

    return jsonify(result)


@worklog_bp.route("/api/worklogs/export/<int:year>/<int:month>", methods=["GET"])
@admin_required
def export_worklog(year, month):
    assistants = Assistant.query.filter(
        Assistant.subject != "퇴직", Assistant.is_admin == False
    ).all()

    wb = Workbook()

    wb.remove(wb.active)

    header_fill = PatternFill(
        start_color="8064A2", end_color="8064A2", fill_type="solid"
    )
    stripe_fill = PatternFill(
        start_color="E4DFEC", end_color="E4DFEC", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(color="FFFFFF", size=14)
    title_fill = PatternFill(
        start_color="1F497D", end_color="1F497D", fill_type="solid"
    )
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_alignment = Alignment(horizontal="center", vertical="center")

    for assistant in assistants:
        sheet = wb.create_sheet(title=f"{assistant.name}")

        title = f"{month}월 급여 {assistant.name} ({assistant.bank_account})"
        sheet.merge_cells("A1:F2")
        title_cell = sheet.cell(row=1, column=1)
        title_cell.value = title
        title_cell.fill = title_fill
        title_cell.font = title_font
        title_cell.border = border
        title_cell.alignment = center_alignment

        sheet.column_dimensions["A"].width = 15  # 날짜
        sheet.column_dimensions["B"].width = 15  # 출근시간
        sheet.column_dimensions["C"].width = 15  # 퇴근시간
        sheet.column_dimensions["D"].width = 15  # 총시간
        sheet.column_dimensions["E"].width = 15  # 시급
        sheet.column_dimensions["F"].width = 15  # 금액

        headers = ["날짜", "출근시간", "퇴근시간", "총시간", "시급", "금액"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=3, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_alignment

        for row in range(4, 31):
            for col in range(1, 7):
                cell = sheet.cell(row=row, column=col)
                cell.border = border
                cell.alignment = center_alignment
                if row % 2 == 0:
                    cell.fill = stripe_fill

        worklogs = (
            WorkLog.query.filter(
                WorkLog.assistant_id == assistant.id,
                extract("year", WorkLog.date) == year,
                extract("month", WorkLog.date) == month,
                WorkLog.checked == True,
            )
            .order_by(WorkLog.date)
            .all()
        )

        current_row = 4
        total_amount = 0
        pointfive = 0
        tax = 0

        for log in worklogs:
            date_str = f"{month}월 {log.date.day}일"

            start_time = log.start_time.strftime("%H:%M")
            end_time = log.end_time.strftime("%H:%M")

            work_hours = f"{int(log.work_hours):02d}:{(log.work_hours % 1 * 60):02.0f}"

            amount = log.work_hours * assistant.salary
            total_amount += amount

            if log.work_hours % 1 == 0.5:
                pointfive += 1

            fake_amount = int(log.work_hours) * assistant.salary

            row_data = [
                date_str,
                start_time,
                end_time,
                work_hours,
                format(assistant.salary, ","),
                format(int(fake_amount), ","),
            ]

            for col, value in enumerate(row_data, 1):
                cell = sheet.cell(row=current_row, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(horizontal="center")

            current_row += 1

        sheet.cell(row=current_row + 5, column=5).value = format(
            assistant.salary // 2, ","
        )
        sheet.cell(row=current_row + 5, column=5).alignment = center_alignment

        sheet.cell(row=current_row + 6, column=4).value = "30분"
        sheet.cell(row=current_row + 6, column=4).alignment = center_alignment
        sheet.cell(row=current_row + 6, column=5).value = pointfive
        sheet.cell(row=current_row + 6, column=5).alignment = center_alignment
        sheet.cell(row=current_row + 6, column=6).value = format(
            pointfive * (assistant.salary // 2), ","
        )
        sheet.cell(row=current_row + 6, column=6).alignment = center_alignment

        sheet.cell(row=current_row + 12, column=6).value = format(
            int(total_amount), ","
        )
        sheet.cell(row=current_row + 12, column=6).alignment = center_alignment

        tax = total_amount * 0.033

        sheet.cell(row=current_row + 15, column=3).value = "3.30%"
        sheet.cell(row=current_row + 15, column=3).alignment = center_alignment
        sheet.cell(row=current_row + 15, column=4).value = tax
        sheet.cell(row=current_row + 15, column=4).alignment = center_alignment
        sheet.cell(row=current_row + 15, column=5).value = "실급여"
        sheet.cell(row=current_row + 15, column=5).alignment = center_alignment
        sheet.cell(row=current_row + 15, column=6).value = format(
            int(total_amount) - int(tax), ","
        )
        sheet.cell(row=current_row + 15, column=6).alignment = center_alignment

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    filename = f"worklog_{year}_{month}.xlsx"

    return send_file(
        excel_file,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )
